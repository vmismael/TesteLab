import streamlit as st
import pandas as pd
import plotly.express as px
import openpyxl
from openpyxl.styles.colors import Color
from datetime import datetime
import re

# ---------------------------------------------------------
# FUNÇÕES AUXILIARES (Para Análise de Medicamentos)
# ---------------------------------------------------------

def get_color_info(cell):
    """Retorna informações detalhadas sobre a cor da célula para debug."""
    fill = cell.fill
    if not fill or not fill.start_color:
        return "Sem Preenchimento", None
    
    color = fill.start_color
    
    if color.type == 'rgb':
        return f"RGB: {color.rgb}", color.rgb
    if color.type == 'theme':
        return f"Tema: {color.theme} (Tint: {color.tint})", 'THEME'
    if color.type == 'indexed':
        return f"Index: {color.indexed}", 'INDEX'
        
    return "Outro", None

def is_green_smart(cell, wb):
    """Verifica se a célula é verde, tentando lidar com Temas e RGB."""
    fill = cell.fill
    if not fill or not fill.start_color:
        return False
    
    color = fill.start_color
    hex_code = None

    if color.type == 'rgb':
        hex_code = color.rgb 
    elif color.type == 'theme':
        if color.theme in [5, 6, 9]: 
            return True 
        return False
        
    if hex_code:
        try:
            if len(hex_code) > 6:
                hex_code = hex_code[2:] 
            if len(hex_code) == 6:
                r = int(hex_code[0:2], 16)
                g = int(hex_code[2:4], 16)
                b = int(hex_code[4:6], 16)
                return g > r and g > b and g > 60
        except:
            pass
    return False

def parse_date(value):
    """Extrai data de strings ou objetos datetime."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', value)
        if match:
            try:
                d, m, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
                if y < 100: y += 2000
                return datetime(y, m, d).date()
            except:
                pass
    return None

# ---------------------------------------------------------
# CONFIGURAÇÃO GERAL DA PÁGINA
# ---------------------------------------------------------
st.set_page_config(page_title="Laboratório Paulista", layout="wide")

st.title("📊 Dashboard Integrado de Gestão")

# ---------------------------------------------------------
# MENU DE NAVEGAÇÃO LATERAL
# ---------------------------------------------------------
st.sidebar.title("Navegação")
pagina_selecionada = st.sidebar.radio(
    "Ir para:",
    [
        "📋 Análise de Coletas", 
        "⚠️ Mapeamento de Riscos", 
        "💊 Análise de Medicamentos",
        "📂 Organizador de Notas"
    ]
)
st.sidebar.markdown("---")

# =========================================================
# PÁGINA 1: ANÁLISE DE COLETAS POR COLABORADOR
# =========================================================
if pagina_selecionada == "📋 Análise de Coletas":
    st.header("Análise de Produtividade por Colaborador")
    st.markdown("Esta ferramenta analisa o arquivo de coletas (CSV) para contabilizar atendimentos.")

    uploaded_file_coletas = st.file_uploader("📂 Carregue o arquivo de Coletas (CSV) aqui", type=["csv"], key="upload_coletas")

    if uploaded_file_coletas:
        try:
            try:
                df_coletas = pd.read_csv(uploaded_file_coletas, sep=";", encoding='utf-8')
            except UnicodeDecodeError:
                uploaded_file_coletas.seek(0)
                df_coletas = pd.read_csv(uploaded_file_coletas, sep=";", encoding='latin1')
            
            if 'Usuário Nome' in df_coletas.columns and 'O.S.' in df_coletas.columns:
                resumo = df_coletas.groupby('Usuário Nome')['O.S.'].nunique().reset_index()
                resumo.columns = ['Colaborador', 'Qtd. Pacientes Atendidos']
                
                resumo_grafico = resumo.sort_values(by='Qtd. Pacientes Atendidos', ascending=True)
                resumo_tabela = resumo.sort_values(by='Qtd. Pacientes Atendidos', ascending=False).reset_index(drop=True)
                
                total_atendimentos = resumo_tabela['Qtd. Pacientes Atendidos'].sum()
                df_total = pd.DataFrame([['TOTAL', total_atendimentos]], columns=['Colaborador', 'Qtd. Pacientes Atendidos'])
                resumo_tabela = pd.concat([resumo_tabela, df_total], ignore_index=True)

                st.subheader("Resumo de Atendimentos")
                col1, col2 = st.columns([1, 2])
                
                with col1:
                    st.write("**Tabela de Dados:**")
                    st.dataframe(
                        resumo_tabela,
                        use_container_width=True,
                        hide_index=True,
                        height=500,
                        column_config={
                            "Qtd. Pacientes Atendidos": st.column_config.NumberColumn(
                                "Qtd. Pacientes",
                                help="Número total de pacientes únicos atendidos",
                                format="%d"
                            )
                        }
                    )
                    
                with col2:
                    st.write("**Gráfico Visual:**")
                    fig = px.bar(
                        resumo_grafico, 
                        x='Qtd. Pacientes Atendidos', 
                        y='Colaborador', 
                        orientation='h', 
                        text_auto=True,
                        title="Pacientes Atendidos por Colaborador"
                    )
                    
                    fig.update_layout(
                        xaxis_title="Quantidade de Pacientes",
                        yaxis_title="Colaborador",
                        showlegend=False,
                        height=500,
                        margin=dict(r=50)
                    )
                    fig.update_traces(textposition='outside', cliponaxis=False)
                    st.plotly_chart(fig, use_container_width=True)

                st.markdown("---")

                st.subheader("🔎 Detalhes por Colaborador")
                st.info("Selecione um colaborador abaixo para ver a lista detalhada.")

                lista_colaboradores = resumo_tabela[resumo_tabela['Colaborador'] != 'TOTAL']['Colaborador'].unique()
                colaborador_selecionado = st.selectbox("Escolha o Colaborador:", lista_colaboradores)

                if colaborador_selecionado:
                    df_filtrado = df_coletas[df_coletas['Usuário Nome'] == colaborador_selecionado].copy()
                    
                    colunas_detalhe = ['Data da Operação', 'O.S.', 'Paciente', 'Paciente Nome', 'Detalhe Descrição']
                    cols_existentes = [c for c in colunas_detalhe if c in df_filtrado.columns]
                    df_detalhe_final = df_filtrado[cols_existentes]
                    
                    df_detalhe_unico = df_detalhe_final.drop_duplicates(subset=['O.S.'])

                    st.write(f"**Pacientes atendidos por: {colaborador_selecionado}**")
                    st.dataframe(
                        df_detalhe_unico, 
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    csv = df_detalhe_unico.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="📥 Baixar detalhes (CSV)",
                        data=csv,
                        file_name=f'detalhes_{colaborador_selecionado}.csv',
                        mime='text/csv',
                    )
            else:
                st.error("O arquivo carregado não possui as colunas 'Usuário Nome' ou 'O.S.'. Verifique se o arquivo está correto.")

        except Exception as e:
            st.error(f"Ocorreu um erro ao processar o arquivo de coletas: {e}")
    else:
        st.info("Por favor, carregue o arquivo CSV de coletas para visualizar os dados.")

# =========================================================
# PÁGINA 2: MAPEAMENTO DE RISCOS
# =========================================================
elif pagina_selecionada == "⚠️ Mapeamento de Riscos":
    st.header("Análise de Riscos Institucionais - Alta e Muito Alta Gravidade")
    st.markdown("""
    Esta ferramenta analisa o arquivo de Mapeamento de Riscos (Excel ou CSV) e filtra eventos classificados como **Alto** ou **Muito Alto**.
    Códigos considerados: `2A`, `3A`, `4A`, `5A`, `3B`, `4B`, `5B`, `5C`.
    """)

    uploaded_file_riscos = st.file_uploader("📂 Carregue seu arquivo Excel ou CSV de Riscos aqui", type=["xlsx", "csv"], key="upload_riscos")

    if uploaded_file_riscos:
        try:
            is_csv = uploaded_file_riscos.name.lower().endswith('.csv')
            
            if is_csv:
                sheet_names = ["Arquivo CSV"]
            else:
                xl = pd.ExcelFile(uploaded_file_riscos)
                sheet_names = [s for s in xl.sheet_names if "Legenda" not in s]
            
            st.sidebar.header("Filtros (Riscos)")
            selected_sheet = st.sidebar.selectbox("Selecione o Setor (Aba):", sheet_names)
            
            months = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
            selected_month = st.sidebar.selectbox("Selecione o Mês:", months)
            
            if st.sidebar.button("🔍 Buscar Riscos", key="btn_buscar_riscos"):
                
                if is_csv:
                    uploaded_file_riscos.seek(0) 
                    df_riscos = pd.read_csv(uploaded_file_riscos, header=None, sep=';', encoding='latin1')
                else:
                    df_riscos = pd.read_excel(uploaded_file_riscos, sheet_name=selected_sheet, header=None)
                
                target_risks = ['2A', '3A', '4A', '5A', '3B', '4B', '5B', '5C']
                
                month_idx = months.index(selected_month)
                content_col_index = 8 + (month_idx * 2)
                risk_col_index = content_col_index + 1
                
                results = []
                
                for index, row in df_riscos.iterrows():
                    first_col = str(row[0])
                    if pd.isna(row[0]) or first_col.strip() in [
                        'FONTE', 'IDENTIFICAÇÃO DO RISCO', 'Identificação do Risco', 
                        'Riscos Institucionais Gerenciados', 
                        'Riscos Institucionais  não Gerenciados/Inventariados', 
                        'C.H.O.R.C.'
                    ]:
                        continue
                    
                    if len(row) > risk_col_index:
                        risk_value = str(row[risk_col_index]).strip().upper()
                        
                        if risk_value in target_risks:
                            results.append({
                                "Identificação do Risco": row[0],
                                "Causa": row[1],
                                f"Conteúdo ({selected_month})": row[content_col_index],
                                "Classificação": risk_value
                            })
                
                if results:
                    st.success(f"Foram encontrados {len(results)} riscos com gravidade Alta/Muito Alta em {selected_sheet} no mês de {selected_month}.")
                    df_results = pd.DataFrame(results)
                    st.dataframe(df_results, use_container_width=True, hide_index=True)
                else:
                    st.info(f"Nenhum risco alto ou muito alto encontrado em {selected_sheet} para {selected_month}.")
                    
        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")
    else:
        st.info("Por favor, carregue o arquivo Excel ou CSV na área acima para começar a análise de riscos.")

# =========================================================
# PÁGINA 3: ANÁLISE DE MEDICAMENTOS
# =========================================================
elif pagina_selecionada == "💊 Análise de Medicamentos":
    st.header("Análise de Atrasos - Medicamentos")
    st.markdown("Verifica células da **Coluna G** (pintadas de verde) e compara com a data de hoje para identificar atrasos na retirada.")

    uploaded_file_med = st.file_uploader("📂 Carregue a planilha (.xlsx)", type=["xlsx"], key="upload_med")

    if uploaded_file_med:
        try:
            wb = openpyxl.load_workbook(uploaded_file_med, data_only=True)
            if 'CONTROLE' in wb.sheetnames:
                ws = wb['CONTROLE']
            else:
                ws = wb.active
                
            hoje = datetime.now().date()
            st.info(f"📅 **Data de Hoje:** {hoje.strftime('%d/%m/%Y')} | Aba analisada: {ws.title}")
            
            atrasados = []
            debug_data = [] 
            
            for i, row in enumerate(ws.iter_rows(min_row=8, min_col=1, max_col=10), start=8):
                cell_date = row[6] 
                cell_name = row[1] 
                cell_med = row[3]  
                
                val_date = cell_date.value
                parsed_date = parse_date(val_date)
                
                e_verde = is_green_smart(cell_date, wb)
                
                color_desc, _ = get_color_info(cell_date)
                debug_data.append({
                    "Linha Excel": i,
                    "Paciente": cell_name.value,
                    "Conteúdo Coluna G": str(val_date),
                    "Data Entendida": parsed_date.strftime('%d/%m/%Y') if parsed_date else "Não detectada",
                    "Cor Detectada": color_desc,
                    "É Verde?": "SIM" if e_verde else "NÃO"
                })
                
                if e_verde and parsed_date:
                    if parsed_date < hoje:
                        atrasados.append({
                            "Linha": i,
                            "Nome do Paciente": cell_name.value,
                            "Medicamento": cell_med.value,
                            "Data Prevista": parsed_date.strftime('%d/%m/%Y'),
                            "Dias de Atraso": (hoje - parsed_date).days
                        })

            if atrasados:
                st.error(f"🚨 **{len(atrasados)} MEDICAMENTOS ATRASADOS ENCONTRADOS!**")
                df_atrasados = pd.DataFrame(atrasados)
                
                try:
                    st.dataframe(df_atrasados.style.background_gradient(cmap="Reds", subset=["Dias de Atraso"]), use_container_width=True)
                except:
                    st.warning("A biblioteca 'matplotlib' não foi encontrada. Exibindo tabela sem gradiente de cores.")
                    st.dataframe(df_atrasados, use_container_width=True)
                
                csv_atraso = df_atrasados.to_csv(index=False).encode('utf-8')
                st.download_button(
                    "📥 Baixar Relatório de Atrasados (CSV)",
                    data=csv_atraso,
                    file_name="medicamentos_atrasados.csv",
                    mime="text/csv"
                )
            else:
                st.success("✅ Nenhum atraso detectado nas células verdes.")
                st.warning("⚠️ Se você vê uma célula verde atrasada e ela não apareceu, verifique o 'Modo Raio-X' abaixo.")

            with st.expander("🔍 MODO RAIO-X (Debug de cores)"):
                st.write("Veja abaixo como o programa leu cada linha. Útil para verificar se a cor verde foi detectada corretamente.")
                df_debug = pd.DataFrame(debug_data)
                st.dataframe(df_debug)

        except Exception as e:
            st.error(f"Erro crítico ao processar o arquivo de medicamentos: {e}")

# =========================================================
# PÁGINA 4: ORGANIZADOR DE NOTAS (ATUALIZADA)
# =========================================================
elif pagina_selecionada == "📂 Organizador de Notas":
    st.header("Organizador de Notas e Arquivos")
    
    uploaded_files_notas = st.file_uploader(
        "Solte as notas em pdf aqui", 
        accept_multiple_files=True, 
        type=['pdf'],
        key="upload_notas"
    )

    if uploaded_files_notas:
        agrupamento = {}
        total_processados = 0
        arquivos_nao_lidos = []

        barra_progresso = st.progress(0)
        
        for i, arquivo in enumerate(uploaded_files_notas):
            nome_arquivo = arquivo.name
            
            # --- LÓGICA DE DETECÇÃO FLEXÍVEL ---
            numero = None
            nome = None
            
            # TENTATIVA 1: Padrão "C 14756 - SINAM.pdf" (Começa com C, depois numero, depois nome)
            match1 = re.search(r"^C\s+(\d+)\s+[-]\s+(.+)\.pdf", nome_arquivo, re.IGNORECASE)
            
            # TENTATIVA 2: Padrão "NOME - 14811.pdf" ou "NOME - 14811 - EXTRA.pdf"
            # Procura qualquer texto no inicio, um traço, e depois numeros
            match2 = re.search(r"^(.+?)\s*[-]\s*(\d+)", nome_arquivo, re.IGNORECASE)

            if match1:
                # No padrão 1, o numero vem primeiro (grupo 1) e o nome depois (grupo 2)
                numero = match1.group(1)
                nome = match1.group(2).upper().strip()
            elif match2:
                # No padrão 2, o nome vem primeiro (grupo 1) e o numero depois (grupo 2)
                nome = match2.group(1).upper().strip()
                numero = match2.group(2)
            
            # -----------------------------------

            if nome and numero:
                if nome not in agrupamento:
                    agrupamento[nome] = []
                agrupamento[nome].append(numero)
                total_processados += 1
            else:
                arquivos_nao_lidos.append(nome_arquivo)
            
            barra_progresso.progress((i + 1) / len(uploaded_files_notas))

        barra_progresso.empty()

        st.success(f"Processamento concluído! {total_processados} arquivos identificados.")
        
        if arquivos_nao_lidos:
            with st.expander(f"⚠️ {len(arquivos_nao_lidos)} arquivos não foram reconhecidos (Ver lista)"):
                st.write("Estes arquivos não seguiram nem o padrão 'C Num - Nome' nem 'Nome - Num':")
                for arq in arquivos_nao_lidos:
                    st.write(f"- {arq}")

        st.write("---")

        if agrupamento:
            col1, col2 = st.columns([1, 2])
            
            with col1:
                st.subheader("Resumo")
                for empresa, lista in agrupamento.items():
                    st.write(f"**{empresa}**: {len(lista)} arquivos")

            with col2:
                st.subheader("Detalhes (Números)")
                for empresa, lista_numeros in agrupamento.items():
                    with st.expander(f"Ver números da {empresa} ({len(lista_numeros)})"):
                        texto_copia = ", ".join(lista_numeros)
                        st.code(texto_copia, language="text")

        else:
            st.warning("Nenhum arquivo compatível foi encontrado.")
